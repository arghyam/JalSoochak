#!/usr/bin/env python3
"""Repair the historical contents of analytics_schema.fact_water_quantity_table.

Why this exists
---------------
Two defects wrote every historical row in that table, and both are fixed in the code
as of the V45 deploy:

1. **Wrong unit.** The column is denominated in litres — that is what every consumer
   reads (total_water_supplied_liters in the CSV/JSON, avgKld = litres/1000 and
   avgLpcd = litres/population in the officer daily report, and the litre-denominated
   efficient-range and performance-score SQL) — but the writers stored the meter's
   native cubic metres. Every stored value is 1000x low.

2. **Wrong baseline.** The day's supply was taken against the *previous calendar day's*
   reading with a COALESCE(..., 0) fallback. On a scheme's first-ever reading, and after
   any gap in submissions, that fallback made the whole cumulative meter index one day's
   supply — values in the millions of m3.

The corrected value for a row is a pure function of the readings:

    (latest reading on D  -  latest reading before D with confirmed_reading > 0) * 1000

which is why this is one uniform recompute-from-readings rather than two targeted
patches. fact_water_quantity_table holds rows from three writers (the pre-go-live
backfill_water_quantity.sql, live ingestion, and the telemetry correction path), and a
recompute lands on the right answer for the two reading-derived ones. Trying to identify
"rows with the baseline bug" and separately "multiply everything by 1000" would
double-apply on any row already written in litres.

The recompute is therefore idempotent, cannot double-apply, and agrees exactly with what
the deployed code now writes — so a reading arriving mid-run is harmless: it fails this
script's value guard and keeps the value the new code just wrote correctly.

The definition of "the corrected value" lives in
backend/analytics-service/src/main/resources/db/scripts/recompute_water_quantity.sql,
which this script reads and WaterQuantityBackfillParityIntegrationTest asserts against
live ingestion. It is not duplicated here.

Run this AFTER deploying analytics-service (Flyway V45 widens the column to BIGINT).
Running it before would leave live ingestion writing cubic metres into a litre table for
the same days.

Case classification
-------------------
Every in-window row is classified into exactly one case. The classification is evaluated
in the order below, so the codes are mutually exclusive:

  A1  no reading row on the date at all               SKIP  — nothing to derive a volume from
  A2  reading row exists, confirmed_reading IS NULL    SKIP  — unconfirmed; becomes derivable
                                                              once someone confirms it (V15
                                                              made the column nullable, so
                                                              this is genuinely reachable)
  B1  first-ever reading for the scheme                apply 0
  B2  earlier readings exist, none usable as a         apply 0
      baseline (all <= 0 or NULL)
  C3  current < previous — meter rollover or           apply 0  LOSSY: the day's real supply
      replacement, clamped by GREATEST(0, ...)                  cannot be recovered
  C4  recomputes above the implausible threshold       SKIP  — a bad reading, not a bad
                                                              recompute. Applying it would
                                                              bake a garbage value in at
                                                              1000x its current size.
  C2  a gap precedes the date (previous_date < D-1)    apply — the multi-day delta
  C1  contiguous previous day, plausible               apply — the ordinary case

A1, A2 and C4 are left holding their pre-repair value, which for a historical row means
they stay in CUBIC METRES. Skipping does not make them correct; it makes them unchanged
and reported. Fix the underlying readings and re-run to pick them up.

Every row the repair declines to touch (A1, A2, C4) is written to an Excel workbook — the
run artefact — together with the full case split, the pre-flight checks, and the
future-dated rows. The applied-but-notable cases (B2, C3, long gaps, duplicates) are
reported as counts on the Summary sheet only; they are all still queryable from
public.fact_water_quantity_recompute by case_code / gap_days / is_latest.

Phase 3 refuses to run when a scheme's dim_scheme rows disagree on
fhtc_count/house_hold_count, because the score formula then has more than one answer per
(scheme, date) and UPDATE ... FROM would pick one arbitrarily — the same replay would give
different scores run to run. --dim-drift-use-latest resolves that deterministically by
scoring each scheme from its most recently updated dim_scheme row.

Phases
------
  0 backup       copy in-window rows of both fact tables to public.*_backup_<suffix>
  1 identify     build public.fact_water_quantity_recompute — the review artefact
  2 water        apply the recompute, chunked by month, value-guarded
  3 performance  replay the daily performance score over the corrected quantities

Modes
-----
  --dry-run   (default) phases 0-1 only; nothing is written to the fact tables
  --execute   phases 0-3
  --verify    report only: what changed, before/after distribution, exceptions
  --rollback  restore both fact tables by id from a phase-0 backup

Usage
-----
  export ANALYTICS_DSN='host=localhost port=5432 dbname=shared_db user=postgres password=...'

  # what would change? review the Excel workbook it writes, especially the skipped cases
  python3 scripts/water_quantity_units_fix.py --dry-run

  # apply it, then check
  python3 scripts/water_quantity_units_fix.py --execute
  python3 scripts/water_quantity_units_fix.py --verify

  # water only, over one month
  python3 scripts/water_quantity_units_fix.py --execute --phase water \
      --start-date 2026-04-01 --end-date 2026-04-30

  # undo
  python3 scripts/water_quantity_units_fix.py --rollback --backup-suffix 20260903_181500

Over a VPN, run this from a host inside the network under tmux, and put TCP keepalives in
the DSN — phase 1 is a single long statement that looks idle at the TCP layer and gets
reaped by NAT/VPN gateways otherwise:

    keepalives=1 keepalives_idle=30 keepalives_interval=10 keepalives_count=5

Pre-create the phase-1 lookup index CONCURRENTLY before the run (see ensure_backfill_index):
the in-script CREATE INDEX takes a SHARE lock that blocks every write to
fact_meter_reading_table for the duration of the build. That is the only step in this
script with meaningful downtime.

After --execute, flush the Redis dashboard keys: SchemeRegularityServiceImpl caches for
24h, so stale 1000x-low values would keep being served for a full day otherwise. Then
VACUUM (ANALYZE) the water table — a full-table UPDATE leaves one dead tuple per row.
"""

from __future__ import annotations

import argparse
import datetime as dt
import logging
import os
import re
import sys
from pathlib import Path

try:
    import psycopg2
    import psycopg2.extras
except ImportError:  # pragma: no cover
    sys.exit("psycopg2 is required:  pip install psycopg2-binary")

LOG = logging.getLogger("water-quantity-fix")

REPO_ROOT = Path(__file__).resolve().parent.parent
RECOMPUTE_SQL_PATH = (REPO_ROOT / "backend" / "analytics-service" / "src" / "main"
                      / "resources" / "db" / "scripts" / "recompute_water_quantity.sql")

WATER_TABLE = "analytics_schema.fact_water_quantity_table"
PERFORMANCE_TABLE = "analytics_schema.fact_scheme_performance_table"
READING_TABLE = "analytics_schema.fact_meter_reading_table"
SCHEME_TABLE = "analytics_schema.dim_scheme_table"
RECOMPUTE_TABLE = "public.fact_water_quantity_recompute"

SAFE_SUFFIX_RE = re.compile(r"^[A-Za-z0-9_]+$")

# Every case the classifier can emit: code -> (skipped_by_design, one-line meaning).
# Ordered as the SQL evaluates them, which is also the order they are reported in.
CASE_CATALOGUE: dict[str, tuple[bool, str]] = {
    "A1": (True, "No reading row on the date at all — no derivable volume"),
    "A2": (True, "Reading row exists but confirmed_reading IS NULL — unconfirmed"),
    "B1": (False, "First-ever reading for the scheme — no baseline, correctly 0"),
    "B2": (False, "Earlier readings exist but none usable as a baseline — correctly 0"),
    "C3": (False, "current < previous (meter rollover/replacement) — clamped to 0, LOSSY"),
    "C4": (True, "Recomputes above the implausible threshold — bad reading, left unchanged"),
    "C2": (False, "A gap precedes the date — delta spans more than one day"),
    "C1": (False, "Contiguous previous day, plausible — the ordinary case"),
}

SKIPPED_CASES = tuple(code for code, (skipped, _) in CASE_CATALOGUE.items() if skipped)

# Mirrors SchemePerformanceSchedulerRepository.insertDailySchemePerformanceScores. Kept as one
# expression so the replay cannot disagree with the scheduler on the thresholds; the "5" is the
# assumed persons-per-household in that query's own comment.
PERFORMANCE_SCORE_CASE = """
    CASE
        WHEN COALESCE(supply.total_water_supplied, 0) <= 0 THEN 0.0
        WHEN COALESCE(supply.total_water_supplied, 0) <
             (
                 COALESCE(ds.fhtc_count, 0) * COALESCE(ds.house_hold_count, 0) * 5
                 * COALESCE(dt.required_lpcd, 0)
             ) THEN 0.5
        ELSE 1.0
    END
"""

# Columns every detail sheet carries, in order. Shared so the workbook reads uniformly and so a
# reviewer can pivot one sheet against another without re-mapping headers.
DETAIL_COLUMNS = [
    "id", "tenant_id", "scheme_id", "date", "case_code", "skipped_by_design",
    "old_qty", "new_qty", "current_reading", "previous_reading", "previous_date",
    "gap_days", "is_latest", "is_future_dated",
]

# (sheet name, WHERE predicate over RECOMPUTE_TABLE, ORDER BY, why it is in the workbook).
# Table-driven so adding a newly discovered edge case is one tuple, not a new function.
DETAIL_SHEETS = [
    ("A - No Usable Reading",
     "case_code IN ('A1','A2')",
     "old_qty DESC, id",
     "SKIPPED. No reading to derive from, so these keep their pre-repair value — which for a "
     "historical row means they are still in CUBIC METRES. A2 becomes repairable once the "
     "reading is confirmed; re-run then."),
    ("C4 - Implausible",
     "case_code = 'C4'",
     "new_qty DESC, id",
     "SKIPPED. The recompute is correct but the underlying confirmed_reading is not — typically "
     "a digit inserted during extraction. Fix the reading in fact_meter_reading_table and re-run; "
     "until then these stay in CUBIC METRES."),
    ("Future Dated",
     "is_future_dated",
     "date DESC, id",
     "APPLIED. These rows are dated after today — a pre-existing data-quality problem this script "
     "neither causes nor fixes. Listed so the repair is not blamed for them."),
]

# B2, C3, long gaps, duplicates and dim-scheme drift are deliberately NOT dumped as detail sheets:
# every one of them is applied (or, for drift, resolved by --dim-drift-use-latest), so a reviewer
# needs the count, not the rows. Their counts stay on the Summary sheet, and every one of them is
# still queryable from RECOMPUTE_TABLE by case_code / gap_days / is_latest.


# --------------------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------------------

def load_recompute_sql() -> str:
    """Reads the canonical recompute definition shared with the parity test."""
    if not RECOMPUTE_SQL_PATH.is_file():
        sys.exit(f"missing recompute definition: {RECOMPUTE_SQL_PATH}\n"
                 "Run this from a checkout of the repository — the SQL is deliberately not "
                 "duplicated in this script.")
    return RECOMPUTE_SQL_PATH.read_text(encoding="utf-8")


def describe_connection(conn) -> str:
    with conn.cursor() as cur:
        cur.execute("SELECT current_database(), current_user, inet_server_addr(), version()")
        db, user, host, version = cur.fetchone()
    return f"{db} as {user} on {host or 'local socket'} — {version.split(',')[0]}"


def resolve_window(conn, start: dt.date | None, end: dt.date | None) -> tuple[dt.date, dt.date]:
    """Window defaults to the full extent of the fact table.

    Deliberately starts before go-live: the pre-go-live rows written by
    backfill_water_quantity.sql are in cubic metres too and need the same conversion, so
    nobody has to pin down the exact go-live date.
    """
    if start and end:
        return start, end
    with conn.cursor() as cur:
        cur.execute(f"SELECT MIN(date), MAX(date) FROM {WATER_TABLE}")
        min_date, max_date = cur.fetchone()
    if min_date is None:
        sys.exit(f"{WATER_TABLE} is empty — nothing to repair, and almost certainly the "
                 "wrong database.")
    return start or min_date, end or max_date


def month_chunks(start: dt.date, end: dt.date):
    """Yields [chunk_start, chunk_end] calendar-month slices covering the window inclusively."""
    cursor = start.replace(day=1)
    while cursor <= end:
        next_month = (cursor.replace(day=28) + dt.timedelta(days=4)).replace(day=1)
        yield max(cursor, start), min(next_month - dt.timedelta(days=1), end)
        cursor = next_month


def scalar(conn, sql: str, params=None):
    with conn.cursor() as cur:
        cur.execute(sql, params)
        row = cur.fetchone()
    return row[0] if row else None


# --------------------------------------------------------------------------------------
# phase 0 — backup
# --------------------------------------------------------------------------------------

def _backup_covers_window(conn, backup_table: str, live_table: str, date_column: str,
                           start: dt.date, end: dt.date) -> bool:
    """True iff every row currently in live_table's window has a matching id in backup_table.

    Id-presence rather than MIN/MAX(date_column) on purpose: a resumed run may ask for a window
    wider than, narrower than, or offset from the one the kept backup was originally taken over,
    and comparing endpoints can't tell "backup is missing dates in the middle" from "there's
    genuinely no data there". Checking that every live row's id is backed up catches both.
    """
    live_count = scalar(conn, f"SELECT COUNT(*) FROM {live_table} WHERE {date_column} BETWEEN %s AND %s",
                        (start, end))
    if not live_count:
        return True
    covered_count = scalar(conn, f"""
        SELECT COUNT(*) FROM {live_table} t
        WHERE t.{date_column} BETWEEN %s AND %s
          AND EXISTS (SELECT 1 FROM {backup_table} b WHERE b.id = t.id)
        """, (start, end))
    return covered_count == live_count


def backup(conn, suffix: str, start: dt.date, end: dt.date, overwrite: bool) -> dict:
    water_backup = f"public.fact_water_quantity_backup_{suffix}"
    performance_backup = f"public.fact_scheme_performance_backup_{suffix}"

    # Resuming an interrupted run means re-running --execute with the same --backup-suffix, and
    # re-taking the backup then would capture the half-repaired table and destroy the only copy of
    # the original values. An existing backup is therefore kept, not replaced — but only if it
    # actually covers this run's window; a --backup-suffix reused across a wider or shifted window
    # would otherwise silently leave the extra rows with no rollback point.
    water_backup_exists = bool(scalar(conn, "SELECT to_regclass(%s)", (water_backup,)))
    if water_backup_exists and not overwrite:
        performance_backup_exists = bool(scalar(conn, "SELECT to_regclass(%s)", (performance_backup,)))
        if not performance_backup_exists:
            sys.exit(f"{water_backup} exists but {performance_backup} does not — backups for a "
                     "suffix must be taken together. Use --overwrite-backup or a different "
                     "--backup-suffix.")
        if not (_backup_covers_window(conn, water_backup, WATER_TABLE, "date", start, end)
                and _backup_covers_window(conn, performance_backup, PERFORMANCE_TABLE,
                                          "last_water_supply_date", start, end)):
            sys.exit(f"{water_backup} / {performance_backup} do not cover every row in "
                     f"{start}..{end} — this --backup-suffix was taken over a different window. "
                     "Use --overwrite-backup to retake it (only safe before any --execute has run "
                     "against this suffix) or pick a fresh --backup-suffix.")
        LOG.info("phase 0: %s already exists and covers %s..%s — kept as the rollback point, "
                 "not re-taken", water_backup, start, end)
    else:
        with conn.cursor() as cur:
            cur.execute(f"DROP TABLE IF EXISTS {water_backup}")
            cur.execute(f"""
                CREATE TABLE {water_backup} AS
                SELECT * FROM {WATER_TABLE} WHERE date BETWEEN %s AND %s
                """, (start, end))
            cur.execute(f"DROP TABLE IF EXISTS {performance_backup}")
            cur.execute(f"""
                CREATE TABLE {performance_backup} AS
                SELECT * FROM {PERFORMANCE_TABLE} WHERE last_water_supply_date BETWEEN %s AND %s
                """, (start, end))

    water_rows = scalar(conn, f"SELECT COUNT(*) FROM {water_backup}")
    water_sum = scalar(conn, f"SELECT COALESCE(SUM(water_quantity), 0) FROM {water_backup}")
    performance_rows = scalar(conn, f"SELECT COUNT(*) FROM {performance_backup}")
    performance_sum = scalar(
        conn, f"SELECT COALESCE(SUM(performance_score), 0) FROM {performance_backup}")

    LOG.info("phase 0: %s — %d row(s), SUM(water_quantity)=%s",
             water_backup, water_rows, water_sum)
    LOG.info("phase 0: %s — %d row(s), SUM(performance_score)=%s",
             performance_backup, performance_rows, performance_sum)
    if water_rows == 0:
        LOG.warning("backed up ZERO water rows for %s..%s — check the window and the database",
                    start, end)
    return {
        "water_backup": water_backup,
        "performance_backup": performance_backup,
        "water_rows": water_rows,
        "water_sum": water_sum,
        "performance_rows": performance_rows,
        "performance_sum": performance_sum,
    }


# --------------------------------------------------------------------------------------
# pre-flight integrity checks
# --------------------------------------------------------------------------------------

def preflight(conn, start: dt.date, end: dt.date) -> list[dict]:
    """Integrity checks for conditions this script's SQL silently tolerates.

    None of these are produced by the repair — they are pre-existing states that change what the
    repair means, or that a reviewer must know about before signing off. Each returns a count; a
    non-zero count is not automatically fatal, so they are reported rather than enforced.
    """
    checks = [
        ("water_quantity IS NULL",
         f"SELECT COUNT(*) FROM {WATER_TABLE} WHERE date BETWEEN %s AND %s "
         "AND water_quantity IS NULL",
         (start, end),
         "Phase 2's value guard (water_quantity = old_qty) is NULL-unsafe, so these rows would "
         "be silently skipped and would not appear in any case bucket. Expect 0; if not, the "
         "case counts will not sum to the total."),
        ("water_quantity < 0",
         f"SELECT COUNT(*) FROM {WATER_TABLE} WHERE date BETWEEN %s AND %s "
         "AND water_quantity < 0",
         (start, end),
         "No writer should ever have produced a negative volume. Expect 0."),
        ("schemes phase 3 will not score",
         f"""SELECT COUNT(*) FROM (
                 SELECT DISTINCT fwq.tenant_id, fwq.scheme_id
                 FROM {WATER_TABLE} fwq
                 WHERE fwq.date BETWEEN %s AND %s
                   AND NOT EXISTS (
                       SELECT 1 FROM {SCHEME_TABLE} ds
                       WHERE ds.tenant_id = fwq.tenant_id AND ds.scheme_id = fwq.scheme_id
                         AND ds.operating_status > 0)
             ) x""",
         (start, end),
         "These schemes have water rows but no dim_scheme row with operating_status > 0, so "
         "phase 3 leaves their performance scores computed against cubic metres."),
        ("dim_scheme attribute drift",
         f"""SELECT COUNT(*) FROM (
                 SELECT tenant_id, scheme_id
                 FROM {SCHEME_TABLE}
                 GROUP BY tenant_id, scheme_id
                 HAVING COUNT(DISTINCT COALESCE(fhtc_count, -1)) > 1
                     OR COUNT(DISTINCT COALESCE(house_hold_count, -1)) > 1
             ) x""",
         None,
         "dim_scheme_table legitimately holds several rows per scheme (V16/V24), but a scheme "
         "whose rows DISAGREE on fhtc_count/house_hold_count yields several different candidate "
         "scores in phase 3, and UPDATE ... FROM picks one arbitrarily. Expect 0; if not, run "
         "scripts/dim_scheme_fanout_diagnostics.sql before phase 3."),
        ("duplicate groups mixing NULL/non-NULL updated_at",
         f"""SELECT COUNT(*) FROM (
                 SELECT tenant_id, scheme_id, date
                 FROM {WATER_TABLE}
                 WHERE date BETWEEN %s AND %s
                 GROUP BY tenant_id, scheme_id, date
                 HAVING COUNT(*) > 1
                    AND COUNT(*) FILTER (WHERE updated_at IS NULL) > 0
                    AND COUNT(*) FILTER (WHERE updated_at IS NOT NULL) > 0
             ) x""",
         (start, end),
         "backfill_water_quantity.sql inserted created_at but no updated_at, and the de-dup "
         "ordering is updated_at DESC (NULLS FIRST), so those rows currently WIN their group. "
         "Phase 2 stamps updated_at = NOW() on every changed row, flipping the winner. "
         "water_quantity is unaffected (the whole group converges on one value), but user_id, "
         "submission_status and outage_reason are read from whichever row wins."),
        ("readings with a tied (date, reading_at)",
         f"""SELECT COUNT(*) FROM (
                 SELECT tenant_id, scheme_id, reading_date
                 FROM {READING_TABLE}
                 WHERE reading_date BETWEEN %s AND %s
                 GROUP BY tenant_id, scheme_id, reading_date
                 HAVING COUNT(*) > COUNT(DISTINCT reading_at)
             ) x""",
         (start, end),
         "'The latest reading on the date' is resolved by reading_at DESC, id DESC. On a tie the "
         "id breaks it — identical to the Java findTopBy...OrderByReadingAtDescIdDesc, so the "
         "repair and live ingestion agree. Informational."),
    ]

    results = []
    for name, sql, params, guidance in checks:
        count = scalar(conn, sql, params)
        results.append({"check": name, "count": count, "guidance": guidance})
        log = LOG.warning if count else LOG.info
        log("preflight: %-46s %s", name, count)
    return results


# --------------------------------------------------------------------------------------
# phase 1 — identify
# --------------------------------------------------------------------------------------

def ensure_backfill_index(conn) -> None:
    """Composite index the recompute's cur/prev LATERAL lookups need to avoid a per-row scan.

    Both lookups in recompute_water_quantity.sql filter by (tenant_id, scheme_id) and order by
    (reading_date, reading_at, id) — the only existing index covers (tenant_id, scheme_id), so
    without this one every row of fact_water_quantity_table drives an unindexed scan of its
    scheme's whole reading history. It also matches the ORDER BY on
    FactMeterReadingRepository's findTopBy...ReadingDateOrderByReadingAtDescIdDesc and
    findLatestBefore, so it keeps paying for itself on the live ingestion path after the backfill.

    NOTE: a plain CREATE INDEX takes a SHARE lock, which blocks every INSERT/UPDATE on
    fact_meter_reading_table until the build finishes — the only step in this script with
    meaningful downtime. On production, build it out-of-band first and let the IF NOT EXISTS
    below turn this into a no-op:

        CREATE INDEX CONCURRENTLY IF NOT EXISTS idx_fact_meter_reading_tenant_scheme_date_lookup
            ON analytics_schema.fact_meter_reading_table
            (tenant_id, scheme_id, reading_date DESC, reading_at DESC, id DESC);
    """
    already_present = bool(scalar(
        conn, "SELECT to_regclass('analytics_schema.idx_fact_meter_reading_tenant_scheme_date_lookup')"))
    if not already_present:
        LOG.warning("prep: building idx_fact_meter_reading_tenant_scheme_date_lookup with a plain "
                    "CREATE INDEX — this holds a SHARE lock and BLOCKS ALL WRITES to %s until it "
                    "completes. On production, cancel and pre-build it CONCURRENTLY instead.",
                    READING_TABLE)
    with conn.cursor() as cur:
        cur.execute(f"""
            CREATE INDEX IF NOT EXISTS idx_fact_meter_reading_tenant_scheme_date_lookup
                ON {READING_TABLE}
                (tenant_id, scheme_id, reading_date DESC, reading_at DESC, id DESC)
            """)
    conn.commit()
    LOG.info("prep: idx_fact_meter_reading_tenant_scheme_date_lookup present on "
             "fact_meter_reading_table")


def identify(conn, start: dt.date, end: dt.date, threshold_litres: int) -> None:
    """Builds public.fact_water_quantity_recompute — the review artefact and phase 2's work list.

    The value columns come verbatim from the canonical recompute; everything added here is
    classification, so the shared SQL that the parity test asserts against stays untouched.

    The two EXISTS probes that separate A1/A2 and B1/B2 sit inside the branches of the outer CASE
    rather than as top-level columns, so Postgres only evaluates them for rows that actually reach
    those branches — a few tens of thousands of index probes rather than one per row.
    """
    recompute_sql = load_recompute_sql()
    with conn.cursor() as cur:
        cur.execute(f"DROP TABLE IF EXISTS {RECOMPUTE_TABLE}")
        cur.execute(f"""
            CREATE TABLE {RECOMPUTE_TABLE} AS
            WITH recompute AS (
            {recompute_sql}
            ),
            windowed AS (
                SELECT * FROM recompute WHERE date BETWEEN %s AND %s
            ),
            coded AS (
                SELECT w.*,
                       CASE
                           WHEN w.current_reading IS NULL THEN
                               CASE WHEN EXISTS (
                                        SELECT 1 FROM {READING_TABLE} mr
                                        WHERE mr.tenant_id = w.tenant_id
                                          AND mr.scheme_id = w.scheme_id
                                          AND mr.reading_date = w.date)
                                    THEN 'A2' ELSE 'A1' END
                           WHEN w.previous_reading IS NULL THEN
                               CASE WHEN EXISTS (
                                        SELECT 1 FROM {READING_TABLE} mr
                                        WHERE mr.tenant_id = w.tenant_id
                                          AND mr.scheme_id = w.scheme_id
                                          AND mr.reading_date < w.date)
                                    THEN 'B2' ELSE 'B1' END
                           WHEN w.current_reading < w.previous_reading THEN 'C3'
                           WHEN w.new_qty > %s THEN 'C4'
                           WHEN w.previous_date < w.date - 1 THEN 'C2'
                           ELSE 'C1'
                       END AS case_code
                FROM windowed w
            )
            SELECT id, tenant_id, scheme_id, date, old_qty, new_qty,
                   current_reading, previous_reading, previous_date, is_latest,
                   case_code,
                   (case_code = ANY(%s)) AS skipped_by_design,
                   (date - previous_date) AS gap_days,
                   (date > CURRENT_DATE) AS is_future_dated,
                   FALSE AS applied
            FROM coded
            """, (start, end, threshold_litres, list(SKIPPED_CASES)))
        cur.execute(f"CREATE UNIQUE INDEX ON {RECOMPUTE_TABLE} (id)")
        cur.execute(f"CREATE INDEX ON {RECOMPUTE_TABLE} (date)")
        cur.execute(f"CREATE INDEX ON {RECOMPUTE_TABLE} (case_code)")


def case_counts(conn) -> list[dict]:
    """Per-case counts, split by what the repair will actually do with each row."""
    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT case_code,
                   COUNT(*),
                   COUNT(*) FILTER (WHERE NOT skipped_by_design
                                      AND new_qty IS NOT NULL AND new_qty <> old_qty),
                   COUNT(*) FILTER (WHERE new_qty IS NOT NULL AND new_qty = old_qty),
                   COUNT(*) FILTER (WHERE skipped_by_design),
                   COUNT(*) FILTER (WHERE skipped_by_design AND old_qty <> 0),
                   COALESCE(SUM(old_qty), 0),
                   COALESCE(SUM(new_qty) FILTER (WHERE NOT skipped_by_design), 0)
            FROM {RECOMPUTE_TABLE}
            GROUP BY case_code
            """)
        rows = {r[0]: r for r in cur.fetchall()}

    counts = []
    for code, (skipped, meaning) in CASE_CATALOGUE.items():
        r = rows.pop(code, None)
        counts.append({
            "case": code,
            "meaning": meaning,
            "action": "SKIP" if skipped else "apply",
            "rows": r[1] if r else 0,
            "will_change": r[2] if r else 0,
            "already_correct": r[3] if r else 0,
            "skipped": r[4] if r else 0,
            "skipped_non_zero": r[5] if r else 0,
            "sum_old_qty": r[6] if r else 0,
            "sum_new_qty_applied": r[7] if r else 0,
        })
    for leftover, r in rows.items():  # defensive: the CASE has an ELSE, so this should stay empty
        counts.append({"case": leftover, "meaning": "UNCLASSIFIED — investigate before applying",
                       "action": "apply", "rows": r[1], "will_change": r[2],
                       "already_correct": r[3], "skipped": r[4], "skipped_non_zero": r[5],
                       "sum_old_qty": r[6], "sum_new_qty_applied": r[7]})
    return counts


def report_cases(conn, counts: list[dict], long_gap_days: int) -> dict:
    """Logs the case split and returns the run-level totals."""
    total = scalar(conn, f"SELECT COUNT(*) FROM {RECOMPUTE_TABLE}")
    LOG.info("phase 1: %s built — %d row(s)", RECOMPUTE_TABLE, total)
    LOG.info("phase 1: case split (action / case / rows / will change / already correct)")
    for c in counts:
        if not c["rows"]:
            continue
        LOG.info("phase 1:   %-5s %-3s %10d rows  %10d change  %8d already correct  — %s",
                 c["action"], c["case"], c["rows"], c["will_change"], c["already_correct"],
                 c["meaning"])

    changing = sum(c["will_change"] for c in counts)
    already_correct = sum(c["already_correct"] for c in counts)
    skipped = sum(c["skipped"] for c in counts)
    skipped_non_zero = sum(c["skipped_non_zero"] for c in counts)
    accounted = changing + already_correct + skipped

    LOG.info("phase 1: %d changing, %d already correct, %d skipped by design (%d of those hold a "
             "non-zero value and stay in CUBIC METRES)",
             changing, already_correct, skipped, skipped_non_zero)
    if accounted != total:
        LOG.warning("phase 1: %d row(s) fall into NO bucket (%d classified vs %d total) — almost "
                    "certainly a NULL water_quantity, which the value guard cannot match. See the "
                    "preflight check.", total - accounted, accounted, total)

    duplicates = scalar(conn, f"SELECT COUNT(*) FROM {RECOMPUTE_TABLE} WHERE NOT is_latest")
    long_gaps = scalar(conn, f"SELECT COUNT(*) FROM {RECOMPUTE_TABLE} "
                             "WHERE case_code IN ('C1','C2') AND gap_days > %s", (long_gap_days,))
    future_dated = scalar(conn, f"SELECT COUNT(*) FROM {RECOMPUTE_TABLE} WHERE is_future_dated")
    if duplicates:
        LOG.info("phase 1:   %d shadow duplicate row(s) behind the latest row of their "
                 "(tenant, scheme, date) — repaired to the same value so none is left in m3",
                 duplicates)
    if long_gaps:
        LOG.warning("phase 1:   %d row(s) attribute a gap of more than %d day(s) to a single date "
                    "— applied, but the daily KLD/LPCD for those dates is distorted",
                    long_gaps, long_gap_days)
    if future_dated:
        LOG.warning("phase 1:   %d row(s) are dated AFTER today — pre-existing bad data, applied "
                    "as derived", future_dated)

    return {
        "total": total,
        "changing": changing,
        "already_correct": already_correct,
        "skipped": skipped,
        "skipped_non_zero": skipped_non_zero,
        "unaccounted": total - accounted,
        "duplicates": duplicates,
        "long_gaps": long_gaps,
        "future_dated": future_dated,
    }


def fetch_detail(conn, predicate: str, order_by: str, limit: int, long_gap_days: int):
    """Rows for one workbook sheet, plus the true (un-truncated) count for that predicate.

    The count is a separate aggregate rather than len(rows) on purpose: a LIMIT-ed fetch reported
    as a count silently understates exactly the lists a reviewer is using to decide whether to
    accept the run.
    """
    params = {"long_gap_days": long_gap_days}
    total = scalar(conn, f"SELECT COUNT(*) FROM {RECOMPUTE_TABLE} WHERE {predicate}", params)
    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT {', '.join(DETAIL_COLUMNS)}
            FROM {RECOMPUTE_TABLE}
            WHERE {predicate}
            ORDER BY {order_by}
            LIMIT {int(limit)}
            """, params)
        rows = cur.fetchall()
    return total, rows


def log_skipped_samples(conn, limit: int) -> None:
    """Prints the skipped rows that hold a non-zero value, keeping the old log shape."""
    total, rows = fetch_detail(conn, "skipped_by_design AND old_qty <> 0", "old_qty DESC, id",
                               limit, 0)
    if not total:
        LOG.info("skipped: none — every skipped row already holds 0")
        return
    LOG.warning("skipped: %d row(s) are left untouched AND hold a non-zero value, so they remain "
                "in CUBIC METRES. Full list in the workbook.", total)
    by_index = {name: i for i, name in enumerate(DETAIL_COLUMNS)}
    for r in rows:
        LOG.warning("  [%s] id=%s tenant=%s scheme=%s date=%s old=%s new=%s "
                    "(reading %s on %s -> %s)",
                    r[by_index["case_code"]], r[by_index["id"]], r[by_index["tenant_id"]],
                    r[by_index["scheme_id"]], r[by_index["date"]], r[by_index["old_qty"]],
                    r[by_index["new_qty"]], r[by_index["previous_reading"]],
                    r[by_index["previous_date"]], r[by_index["current_reading"]])
    if total > limit:
        LOG.warning("  ... %d more; see the workbook or: SELECT * FROM %s WHERE "
                    "skipped_by_design AND old_qty <> 0", total - limit, RECOMPUTE_TABLE)


# --------------------------------------------------------------------------------------
# phase 2 — apply
# --------------------------------------------------------------------------------------

def apply_water(conn, start: dt.date, end: dt.date) -> int:
    """Applies the recompute one calendar month at a time, committing each chunk.

    The UPDATE is guarded on the value this script read (water_quantity = old_qty), which is
    what makes it resumable and safe to run against a live system: a row the new code rewrote
    since phase 1 simply fails the guard and keeps the value the new code wrote.

    NOT skipped_by_design excludes A1/A2 (no derivable value) and C4 (derivable but the source
    reading is garbage). Because the case is a function of (tenant, scheme, date) only, every
    duplicate row of a group is classified identically — a group is never half-applied.
    """
    total = 0
    for chunk_start, chunk_end in month_chunks(start, end):
        with conn.cursor() as cur:
            cur.execute(f"""
                UPDATE {WATER_TABLE} AS fwq
                SET water_quantity = r.new_qty,
                    updated_at = NOW()
                FROM {RECOMPUTE_TABLE} r
                WHERE fwq.id = r.id
                  AND r.date BETWEEN %s AND %s
                  AND NOT r.skipped_by_design
                  AND r.new_qty IS NOT NULL
                  AND r.new_qty <> r.old_qty
                  AND fwq.water_quantity = r.old_qty
                """, (chunk_start, chunk_end))
            updated = cur.rowcount
            cur.execute(f"""
                UPDATE {RECOMPUTE_TABLE} r
                SET applied = TRUE
                FROM {WATER_TABLE} fwq
                WHERE fwq.id = r.id
                  AND r.date BETWEEN %s AND %s
                  AND NOT r.skipped_by_design
                  AND r.new_qty IS NOT NULL
                  AND fwq.water_quantity = r.new_qty
                """, (chunk_start, chunk_end))
        conn.commit()
        total += updated
        LOG.info("phase 2: %s..%s — %d row(s) updated", chunk_start, chunk_end, updated)

    stranded = scalar(conn, f"""
        SELECT COUNT(*) FROM {RECOMPUTE_TABLE}
        WHERE NOT skipped_by_design AND new_qty IS NOT NULL AND new_qty <> old_qty AND NOT applied
        """)
    if stranded:
        LOG.warning("phase 2: %d row(s) failed the value guard — rewritten by live ingestion "
                    "since phase 1, so they already hold a correctly derived value. Re-run "
                    "--dry-run to confirm.", stranded)
    LOG.info("phase 2: %d row(s) updated in total", total)
    return total


# --------------------------------------------------------------------------------------
# phase 3 — performance
# --------------------------------------------------------------------------------------

def dim_scheme_drift(conn) -> list[tuple]:
    """Schemes whose several dim_scheme rows disagree on the attributes the score formula reads.

    dim_scheme_table legitimately holds one row per (scheme x village x sub-division) since
    V16/V24, but fhtc_count and house_hold_count describe the SCHEME, so every row of a scheme
    must carry the same values. When they don't, the scores CTE produces several different
    candidate scores for one (tenant, scheme, date) and UPDATE ... FROM picks one arbitrarily —
    the same replay then yields a different score run to run. Measured on a fixture: a scheme
    with fhtc_count 10 vs 99 flipped 0.5 -> 1.0 -> 0.5 across three identical runs.

    The production scheduler has the same nondeterminism, so this is not introduced here; but a
    repair that writes an arbitrary score to history is worse than one that stops and says so.
    """
    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT tenant_id, scheme_id,
                   COUNT(*) AS dim_rows,
                   ARRAY_AGG(DISTINCT fhtc_count) AS fhtc_counts,
                   ARRAY_AGG(DISTINCT house_hold_count) AS household_counts
            FROM {SCHEME_TABLE}
            GROUP BY tenant_id, scheme_id
            HAVING COUNT(DISTINCT COALESCE(fhtc_count, -1)) > 1
                OR COUNT(DISTINCT COALESCE(house_hold_count, -1)) > 1
            ORDER BY tenant_id, scheme_id
            """)
        return cur.fetchall()


def apply_performance(conn, start: dt.date, end: dt.date, dim_drift_use_latest: bool) -> int:
    """Replays the daily performance score over the corrected quantities.

    dim_drift_use_latest collapses dim_scheme_table to one row per (tenant, scheme) — the most
    recently updated one — before the join, which is what makes the replay deterministic when a
    scheme's rows disagree. Note it also changes what operating_status > 0 means: "the scheme's
    current row says it is operating", rather than the raw scheduler's "some row of the scheme
    says so". That is the more defensible reading of a slowly-changing dimension, but it is a
    deliberate divergence from the scheduler, so it only happens when the flag asks for it.

    Updates existing rows only. The scheduler's own INSERT is guarded by NOT EXISTS, so it
    would no-op over history; and creating scores for days the scheduler never ran would
    invent history rather than repair it. Days with no performance row are counted and
    reported, not filled in.

    Reads the water table rather than the recompute table, so rows this run skipped still
    contribute their (unrepaired, cubic-metre) value to the day's supply — exactly as the
    scheduler would see them. Those (scheme, date) pairs are listed in the workbook.
    """
    scheme_source = f"""(
                        SELECT DISTINCT ON (tenant_id, scheme_id) *
                        FROM {SCHEME_TABLE}
                        ORDER BY tenant_id, scheme_id,
                                 COALESCE(updated_at, created_at) DESC NULLS LAST, id DESC
                    )""" if dim_drift_use_latest else SCHEME_TABLE

    total = 0
    for chunk_start, chunk_end in month_chunks(start, end):
        with conn.cursor() as cur:
            cur.execute(f"""
                WITH latest AS (
                    SELECT DISTINCT ON (fwq.tenant_id, fwq.scheme_id, fwq.date)
                           fwq.tenant_id, fwq.scheme_id, fwq.date, fwq.water_quantity
                    FROM {WATER_TABLE} fwq
                    WHERE fwq.date BETWEEN %s AND %s
                    ORDER BY fwq.tenant_id, fwq.scheme_id, fwq.date,
                             fwq.updated_at DESC, fwq.id DESC
                ),
                supply AS (
                    SELECT tenant_id, scheme_id, date,
                           SUM(water_quantity) AS total_water_supplied
                    FROM latest
                    GROUP BY tenant_id, scheme_id, date
                ),
                scores AS (
                    SELECT ds.tenant_id,
                           ds.scheme_id,
                           supply.date,
                           {PERFORMANCE_SCORE_CASE} AS performance_score
                    FROM {scheme_source} ds
                    JOIN analytics_schema.dim_tenant_table dt
                      ON dt.tenant_id = ds.tenant_id
                    JOIN supply
                      ON supply.tenant_id = ds.tenant_id
                     AND supply.scheme_id = ds.scheme_id
                    WHERE ds.operating_status > 0
                )
                UPDATE {PERFORMANCE_TABLE} fp
                SET performance_score = scores.performance_score,
                    updated_at = NOW()
                FROM scores
                WHERE fp.tenant_id = scores.tenant_id
                  AND fp.scheme_id = scores.scheme_id
                  AND fp.last_water_supply_date = scores.date
                  AND fp.performance_score IS DISTINCT FROM scores.performance_score
                """, (chunk_start, chunk_end))
            updated = cur.rowcount
        conn.commit()
        total += updated
        LOG.info("phase 3: %s..%s — %d score(s) updated", chunk_start, chunk_end, updated)

    missing = scalar(conn, f"""
        SELECT COUNT(*)
        FROM (
            SELECT DISTINCT fwq.tenant_id, fwq.scheme_id, fwq.date
            FROM {WATER_TABLE} fwq
            JOIN {SCHEME_TABLE} ds
              ON ds.tenant_id = fwq.tenant_id AND ds.scheme_id = fwq.scheme_id
             AND ds.operating_status > 0
            WHERE fwq.date BETWEEN %s AND %s
              AND NOT EXISTS (
                  SELECT 1 FROM {PERFORMANCE_TABLE} fp
                  WHERE fp.tenant_id = fwq.tenant_id
                    AND fp.scheme_id = fwq.scheme_id
                    AND fp.last_water_supply_date = fwq.date
              )
        ) gaps
        """, (start, end))
    if missing:
        LOG.info("phase 3: %d (scheme, date) pair(s) have water but no performance row — the "
                 "scheduler never scored those days. Not created here.", missing)
    LOG.info("phase 3: %d score(s) updated in total", total)
    return total


# --------------------------------------------------------------------------------------
# phase 4 — verify
# --------------------------------------------------------------------------------------

def verify(conn, start: dt.date, end: dt.date) -> tuple[int, dict]:
    """Reports on an applied (or partially applied) repair and returns the still-unrepaired count.

    at_old counts only rows the repair was SUPPOSED to change, so it is the caller's signal for a
    non-zero exit status. Rows skipped by design sit at their old value permanently and must not
    make a successful run look like a failure — they are counted and reported separately.
    """
    if not scalar(conn, "SELECT to_regclass(%s)", (RECOMPUTE_TABLE,)):
        sys.exit(f"{RECOMPUTE_TABLE} does not exist — run --dry-run or --execute first.")
    # A table left behind by an older revision of this script has no classification columns, and
    # every query below would fail halfway through on a missing column instead of up front.
    if not scalar(conn, """
            SELECT COUNT(*) = 2 FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = 'fact_water_quantity_recompute'
              AND column_name IN ('case_code', 'skipped_by_design')
            """):
        sys.exit(f"{RECOMPUTE_TABLE} predates the case classification — re-run --dry-run to "
                 "rebuild it before verifying.")

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT COUNT(*),
                   COUNT(*) FILTER (WHERE fwq.water_quantity = r.new_qty),
                   COUNT(*) FILTER (WHERE fwq.water_quantity = r.old_qty
                                      AND r.old_qty <> r.new_qty),
                   COUNT(*) FILTER (WHERE fwq.water_quantity NOT IN (r.old_qty, r.new_qty))
            FROM {RECOMPUTE_TABLE} r
            JOIN {WATER_TABLE} fwq ON fwq.id = r.id
            WHERE r.new_qty IS NOT NULL AND NOT r.skipped_by_design
            """)
        total, at_new, at_old, elsewhere = cur.fetchone()
    LOG.info("verify: %d applicable row(s) — %d at the corrected value, %d still at the old "
             "value, %d at neither (rewritten by live ingestion since)",
             total, at_new, at_old, elsewhere)
    if at_old:
        LOG.warning("verify: %d row(s) were NOT applied — re-run --execute", at_old)

    skipped = scalar(conn, f"SELECT COUNT(*) FROM {RECOMPUTE_TABLE} WHERE skipped_by_design")
    skipped_non_zero = scalar(conn, f"SELECT COUNT(*) FROM {RECOMPUTE_TABLE} "
                                    "WHERE skipped_by_design AND old_qty <> 0")
    LOG.info("verify: %d row(s) skipped by design (%s); %d of them hold a non-zero value and "
             "remain in CUBIC METRES", skipped, "/".join(SKIPPED_CASES), skipped_non_zero)

    with conn.cursor() as cur:
        cur.execute(f"""
            SELECT MIN(old_qty), MAX(old_qty), ROUND(AVG(old_qty)),
                   MIN(new_qty), MAX(new_qty), ROUND(AVG(new_qty))
            FROM {RECOMPUTE_TABLE}
            WHERE new_qty IS NOT NULL AND NOT skipped_by_design
            """)
        row = cur.fetchone()
    # Distribution as of whenever phase 1 last ran. Inside --execute that is the pre-repair data, so
    # the two lines are genuinely before/after. A standalone --verify long after the fact reads a
    # recompute table built from already-corrected rows, where the two lines simply agree — which is
    # itself the confirmation that nothing has drifted since.
    LOG.info("verify: old_qty  min=%s max=%s avg=%s", row[0], row[1], row[2])
    LOG.info("verify: new_qty  min=%s max=%s avg=%s", row[3], row[4], row[5])

    log_skipped_samples(conn, limit=20)

    supplied_before = scalar(conn, f"""
        SELECT COALESCE(SUM(old_qty), 0) FROM {RECOMPUTE_TABLE}
        WHERE is_latest AND new_qty IS NOT NULL
        """)
    supplied_after = scalar(conn, f"""
        SELECT COALESCE(SUM(CASE WHEN skipped_by_design THEN old_qty ELSE new_qty END), 0)
        FROM {RECOMPUTE_TABLE}
        WHERE is_latest AND new_qty IS NOT NULL
        """)
    LOG.info("verify: total over %s..%s — %s stored before, %s stored after (skipped rows counted "
             "at their unchanged value)", start, end, supplied_before, supplied_after)
    LOG.info("verify: remember to flush the Redis dashboard keys (24h TTL) and to "
             "VACUUM (ANALYZE) %s", WATER_TABLE)

    return at_old, {
        "applicable": total, "at_new": at_new, "at_old": at_old, "elsewhere": elsewhere,
        "skipped": skipped, "skipped_non_zero": skipped_non_zero,
        "old_min": row[0], "old_max": row[1], "old_avg": row[2],
        "new_min": row[3], "new_max": row[4], "new_avg": row[5],
        "stored_before": supplied_before, "stored_after": supplied_after,
    }


# --------------------------------------------------------------------------------------
# workbook — the run artefact
# --------------------------------------------------------------------------------------

def write_workbook(conn, path: Path, meta: dict, counts: list[dict], totals: dict,
                   preflight_rows: list[dict], long_gap_days: int, max_rows: int) -> None:
    """Writes the run's Excel artefact: the count split, the preflight checks, and every
    skipped / lossy / ambiguous row.

    write_only keeps the workbook streaming rather than materialising every sheet in memory —
    the long-gap and duplicate sheets can each run to tens of thousands of rows.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font
    except ImportError:
        LOG.error("openpyxl is not installed, so no workbook was written: pip install openpyxl "
                  "(or pass --no-report). Everything in it is still queryable from %s.",
                  RECOMPUTE_TABLE)
        return

    wb = Workbook(write_only=True)
    bold = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    def header(ws, names):
        cells = []
        for name in names:
            c = WriteOnlyCell(ws, value=name)
            c.font = bold
            cells.append(c)
        ws.append(cells)

    # --- Summary -----------------------------------------------------------------------
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["I"].width = 80
    header(ws, ["Water quantity repair — run summary"])
    ws.append([])
    for key, value in meta.items():
        ws.append([key, str(value)])
    ws.append([])

    header(ws, ["Case", "Action", "Rows", "Will change", "Already correct",
                "Skipped", "Skipped & non-zero", "SUM(old_qty)", "SUM(new_qty) applied",
                "Meaning"])
    for c in counts:
        ws.append([c["case"], c["action"], c["rows"], c["will_change"], c["already_correct"],
                   c["skipped"], c["skipped_non_zero"], c["sum_old_qty"],
                   c["sum_new_qty_applied"], c["meaning"]])
    ws.append([])
    header(ws, ["Total", "", "Rows", "Will change", "Already correct", "Skipped",
                "Skipped & non-zero"])
    ws.append(["ALL", "", totals["total"], totals["changing"], totals["already_correct"],
               totals["skipped"], totals["skipped_non_zero"]])
    ws.append([])
    header(ws, ["Cross-cutting flag", "Rows", "Note"])
    ws.append(["Unclassified (no bucket)", totals["unaccounted"],
               "Must be 0. Anything here is invisible to the value guard — investigate."])
    ws.append(["Shadow duplicates", totals["duplicates"],
               "Repaired to the same value as their group's winner."])
    ws.append([f"Gaps longer than {long_gap_days} day(s)", totals["long_gaps"],
               "Applied. A multi-day delta attributed to one date."])
    ws.append(["Future-dated rows", totals["future_dated"],
               "Applied. Dated after today — pre-existing bad data."])

    # --- Preflight ---------------------------------------------------------------------
    ws = wb.create_sheet("Preflight")
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 110
    header(ws, ["Check", "Count", "What a non-zero count means"])
    for row in preflight_rows:
        c = WriteOnlyCell(ws, value=row["guidance"])
        c.alignment = wrap
        ws.append([row["check"], row["count"], c])

    # --- One sheet per reportable case -------------------------------------------------
    for name, predicate, order_by, note in DETAIL_SHEETS:
        total, rows = fetch_detail(conn, predicate, order_by, max_rows, long_gap_days)
        ws = wb.create_sheet(name[:31])
        ws.freeze_panes = "A4"
        ws.column_dimensions["A"].width = 14
        note_cell = WriteOnlyCell(ws, value=note)
        note_cell.alignment = wrap
        ws.append([note_cell])
        truncated = f"  (showing the first {max_rows:,} — raise --max-report-rows)" if total > max_rows else ""
        ws.append([f"{total:,} row(s) match{truncated}"])
        header(ws, DETAIL_COLUMNS)
        for r in rows:
            ws.append(list(r))
        LOG.info("report: %-26s %8d row(s)%s", name, total,
                 " [TRUNCATED]" if total > max_rows else "")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    LOG.info("report: workbook written to %s", path)


# --------------------------------------------------------------------------------------
# rollback
# --------------------------------------------------------------------------------------

def rollback(conn, suffix: str) -> None:
    water_backup = f"public.fact_water_quantity_backup_{suffix}"
    performance_backup = f"public.fact_scheme_performance_backup_{suffix}"
    for table in (water_backup, performance_backup):
        if not scalar(conn, "SELECT to_regclass(%s)", (table,)):
            sys.exit(f"{table} does not exist — check --backup-suffix.")

    with conn.cursor() as cur:
        cur.execute(f"""
            UPDATE {WATER_TABLE} fwq
            SET water_quantity = b.water_quantity,
                updated_at = b.updated_at
            FROM {water_backup} b
            WHERE fwq.id = b.id
              AND fwq.water_quantity IS DISTINCT FROM b.water_quantity
            """)
        water_restored = cur.rowcount
        cur.execute(f"""
            UPDATE {PERFORMANCE_TABLE} fp
            SET performance_score = b.performance_score,
                updated_at = b.updated_at
            FROM {performance_backup} b
            WHERE fp.id = b.id
              AND fp.performance_score IS DISTINCT FROM b.performance_score
            """)
        performance_restored = cur.rowcount
    conn.commit()
    LOG.info("rollback: restored %d water row(s) and %d performance row(s) from %s",
             water_restored, performance_restored, suffix)


# --------------------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------------------

def iso_date(value: str) -> dt.date:
    try:
        return dt.date.fromisoformat(value)
    except ValueError:
        raise argparse.ArgumentTypeError(f"expected YYYY-MM-DD, got {value!r}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true",
                      help="(default) back up and identify; write nothing to the fact tables")
    mode.add_argument("--execute", action="store_true", help="apply the repair")
    mode.add_argument("--verify", action="store_true",
                      help="report on an already-applied repair")
    mode.add_argument("--rollback", action="store_true",
                      help="restore from a phase-0 backup; needs --backup-suffix")
    parser.add_argument("--phase", choices=("water", "performance", "all"), default="all",
                        help="performance must run after water and over the same window")
    parser.add_argument("--dsn", default=os.environ.get("ANALYTICS_DSN"),
                        help="libpq DSN; defaults to $ANALYTICS_DSN")
    parser.add_argument("--start-date", type=iso_date, default=None,
                        help="defaults to MIN(date) of the fact table")
    parser.add_argument("--end-date", type=iso_date, default=None,
                        help="defaults to MAX(date) of the fact table")
    parser.add_argument("--backup-suffix", default=None,
                        help="names the phase-0 backup tables; defaults to a UTC timestamp")
    parser.add_argument("--overwrite-backup", action="store_true",
                        help="re-take a phase-0 backup that already exists, discarding the "
                             "original values it holds")
    parser.add_argument("--implausible-daily-cubic-metres", type=int, default=100_000,
                        help="rows recomputing above this are classified C4 and SKIPPED; "
                             "matches the service's warning threshold")
    parser.add_argument("--long-gap-days", type=int, default=90,
                        help="rows whose delta spans more than this many days are reported "
                             "(still applied)")
    parser.add_argument("--report-path", type=Path, default=None,
                        help="workbook path; defaults to ./water_quantity_fix_<suffix>.xlsx")
    parser.add_argument("--no-report", action="store_true", help="skip the Excel workbook")
    parser.add_argument("--max-report-rows", type=int, default=50_000,
                        help="row cap per workbook sheet; the true count is always reported")
    parser.add_argument("--skip-preflight", action="store_true",
                        help="skip the integrity checks (they scan the reading table)")
    parser.add_argument("--dim-drift-use-latest", action="store_true",
                        help="resolve dim_scheme drift by using each scheme's most recently "
                             "updated row (COALESCE(updated_at, created_at) DESC, id DESC); "
                             "without this, phase 3 refuses to run on drift and exits 2")
    parser.add_argument("--statement-timeout", default="15min",
                        help="server-side statement_timeout for every statement")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO,
                        format="%(levelname)s %(message)s")

    if not args.dsn:
        sys.exit("no DSN: pass --dsn or set ANALYTICS_DSN")

    suffix = args.backup_suffix or dt.datetime.now(dt.timezone.utc).strftime("%Y%m%d_%H%M%S")
    if not SAFE_SUFFIX_RE.match(suffix):
        sys.exit(f"--backup-suffix must be alphanumeric/underscore, got {suffix!r}")
    threshold_litres = args.implausible_daily_cubic_metres * 1000
    report_path = args.report_path or Path(f"water_quantity_fix_{suffix}.xlsx")

    conn = psycopg2.connect(args.dsn)
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            cur.execute("SET statement_timeout = %s", (args.statement_timeout,))
        connection_description = describe_connection(conn)
        LOG.info("connected: %s", connection_description)

        if args.rollback:
            if not args.backup_suffix:
                sys.exit("--rollback needs --backup-suffix (the timestamp phase 0 printed)")
            rollback(conn, suffix)
            return 0

        start, end = resolve_window(conn, args.start_date, args.end_date)
        mode_name = "verify" if args.verify else ("execute" if args.execute else "dry-run")
        LOG.info("window: %s..%s, phase=%s, mode=%s", start, end, args.phase, mode_name)
        LOG.info("C4 threshold: %d L/day (%d m3/day) — rows above this are SKIPPED, not applied",
                 threshold_litres, args.implausible_daily_cubic_metres)

        meta = {
            "generated_at (UTC)": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
            "connection": connection_description,
            "mode": mode_name,
            "phase": args.phase,
            "window start": start,
            "window end": end,
            "backup suffix": suffix,
            "C4 threshold (L/day)": threshold_litres,
            "C4 threshold (m3/day)": args.implausible_daily_cubic_metres,
            "long gap threshold (days)": args.long_gap_days,
            "recompute artefact": RECOMPUTE_TABLE,
        }

        if args.verify:
            at_old, verify_stats = verify(conn, start, end)
            meta.update({f"verify: {k}": v for k, v in verify_stats.items()})
            if not args.no_report:
                counts = case_counts(conn)
                totals = report_cases(conn, counts, args.long_gap_days)
                write_workbook(conn, report_path, meta, counts, totals,
                               [] if args.skip_preflight else preflight(conn, start, end),
                               args.long_gap_days, args.max_report_rows)
            return 1 if at_old else 0

        backup_stats = backup(conn, suffix, start, end, args.overwrite_backup)
        conn.commit()
        meta.update({f"backup: {k}": v for k, v in backup_stats.items()})

        preflight_rows = [] if args.skip_preflight else preflight(conn, start, end)

        ensure_backfill_index(conn)

        identify(conn, start, end, threshold_litres)
        conn.commit()
        counts = case_counts(conn)
        totals = report_cases(conn, counts, args.long_gap_days)
        log_skipped_samples(conn, limit=20)

        if not args.execute:
            if not args.no_report:
                write_workbook(conn, report_path, meta, counts, totals, preflight_rows,
                               args.long_gap_days, args.max_report_rows)
            LOG.info("dry run — nothing written to %s. Review the workbook and %s, then re-run "
                     "with --execute.", WATER_TABLE, RECOMPUTE_TABLE)
            LOG.info("dry run: %d row(s) would change, %d left skipped by design",
                     totals["changing"], totals["skipped"])
            return 0

        if args.phase in ("water", "all"):
            apply_water(conn, start, end)
        if args.phase in ("performance", "all"):
            drift = dim_scheme_drift(conn)
            if drift and not args.dim_drift_use_latest:
                LOG.error("phase 3 REFUSED: %d scheme(s) have dim_scheme rows that disagree on "
                          "fhtc_count/house_hold_count, so the score formula has several answers "
                          "per (scheme, date) and UPDATE ... FROM would pick one arbitrarily — "
                          "the same replay gives different scores run to run. Phase 2 is already "
                          "applied and is unaffected.", len(drift))
                for tenant_id, scheme_id, dim_rows, fhtc, households in drift[:20]:
                    LOG.error("  tenant=%s scheme=%s dim_rows=%s fhtc_count=%s house_hold_count=%s",
                              tenant_id, scheme_id, dim_rows, fhtc, households)
                if len(drift) > 20:
                    LOG.error("  ... %d more; full list: %s", len(drift) - 20,
                              "scripts/dim_scheme_fanout_diagnostics.sql")
                LOG.error("Either repair dim_scheme_table (see "
                          "scripts/dim_scheme_fanout_diagnostics.sql) and re-run with "
                          "--phase performance, or pass --dim-drift-use-latest to score each "
                          "drifted scheme from its most recently updated row.")
                # Phase 2 has already applied at this point, so its artefact is still owed.
                if not args.no_report:
                    write_workbook(conn, report_path, meta, counts, totals, preflight_rows,
                                   args.long_gap_days, args.max_report_rows)
                return 2
            if drift:
                LOG.warning("phase 3: %d scheme(s) have drifted dim_scheme rows — scoring each "
                            "from its most recently updated row "
                            "(COALESCE(updated_at, created_at) DESC, id DESC) because "
                            "--dim-drift-use-latest was passed", len(drift))
            apply_performance(conn, start, end, args.dim_drift_use_latest)

        at_old, verify_stats = verify(conn, start, end)
        meta.update({f"verify: {k}": v for k, v in verify_stats.items()})
        if not args.no_report:
            write_workbook(conn, report_path, meta, counts, totals, preflight_rows,
                           args.long_gap_days, args.max_report_rows)
        LOG.info("done. Backup suffix for --rollback: %s", suffix)
        return 1 if at_old else 0
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
